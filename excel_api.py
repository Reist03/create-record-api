from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import load_workbook
import io
import json
import os
from copy import deepcopy

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://kawatsu624.hiho.jp",
        "http://kawatsu624.hiho.jp",
        "http://localhost:5173",
        "http://127.0.0.1:5173",
    ],
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

TEMPLATE_PATH = os.path.join(
    BASE_DIR,
    "templates",
    "08_ﾓﾆﾀﾘﾝｸﾞ結果印刷_ACE既定（モニタリング）_202404.xlsx"
)


class ExcelRequest(BaseModel):
    summary_json: str | None = None
    text: str | None = None
    transcript: str | None = None


DEFAULT_REPORT = {
    "実施日": "",
    "前回実施日": "",
    "次回予定日": "",

    "利用者名": "",
    "利用者名カナ": "",
    "出力氏名": "",

    "性別": "",
    "生年月日": "",
    "年齢": "",
    "介護度": "",
    "認定開始日": "",
    "認定終了日": "",

    "住所": "",
    "住所1": "",
    "住所2": "",
    "住所3": "",
    "電話番号": "",
    "電話番号1": "",

    "担当名": "",
    "ケアマネ姓名": "",

    "お話伺った人1": "",
    "お話伺った人2": "",
    "お話伺った人3": "",
    "お話伺った人その他": "",

    "確認方法1": "",
    "確認方法2": "",

    "専門相談員による結果": "",

    "福祉用具利用目標": [],
    "商品一覧": [],

    "身体状況の変化1": "",
    "身体状況の変化2": "",
    "身体状況の変化備考": "",

    "ご家族状況の変化1": "",
    "ご家族状況の変化2": "",
    "ご家族状況の変化備考": "",

    "お気持ちの変化1": "",
    "お気持ちの変化2": "",
    "お気持ちの変化備考": "",

    "生活状況の変化1": "",
    "生活状況の変化2": "",
    "生活状況の変化備考": "",

    "見直しの必要性1": "",
    "見直しの必要性2": ""
}


DEFAULT_GOAL = {
    "目標": "",
    "達成度": "",
    "備考": ""
}


DEFAULT_PRODUCT = {
    "対応理由記号": "",
    "選択制対象区分": "",

    "サービス名": "",
    "利用開始日": "",
    "商品名": "",

    "使用状況の問題1": "",
    "点検結果1": "",
    "今後の方針1": "",

    "使用状況の問題2": "",
    "点検結果2": "",
    "今後の方針2": "",

    "モニタリング備考": ""
}


def merge_dict(defaults: dict, actual: dict) -> dict:
    result = deepcopy(defaults)

    if not isinstance(actual, dict):
        return result

    for key, value in actual.items():
        result[key] = value

    return result


def normalize_report(raw: dict) -> dict:
    report = merge_dict(DEFAULT_REPORT, raw if isinstance(raw, dict) else {})

    goals = report.get("福祉用具利用目標")
    if not isinstance(goals, list):
        goals = []

    normalized_goals = []
    for g in goals[:4]:
        normalized_goals.append(
            merge_dict(DEFAULT_GOAL, g if isinstance(g, dict) else {})
        )
    report["福祉用具利用目標"] = normalized_goals

    products = report.get("商品一覧")
    if not isinstance(products, list):
        products = []

    normalized_products = []
    for p in products[:8]:
        normalized_products.append(
            merge_dict(DEFAULT_PRODUCT, p if isinstance(p, dict) else {})
        )
    report["商品一覧"] = normalized_products

    for key, value in list(report.items()):
        if value is None:
            report[key] = ""

    return report


def safe_json(text: str) -> dict:
    try:
        raw = json.loads(text)
        return normalize_report(raw)
    except Exception:
        fallback = deepcopy(DEFAULT_REPORT)
        fallback["専門相談員による結果"] = text
        return fallback


# 修正版
def set_if_exists(ws, cell_ref: str, value):
    value = "" if value is None else value

    cell = ws[cell_ref]

    # 結合セル対応
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            top_left_cell = ws.cell(
                row=merged_range.min_row,
                column=merged_range.min_col
            )
            top_left_cell.value = value
            return

    cell.value = value


def norm(value) -> str:
    return "" if value is None else str(value).strip()


def mark_choice(value, expected_values) -> str:
    value = norm(value)
    expected_values = [str(v).strip() for v in expected_values]
    return "〇" if value in expected_values else ""


def mark_achievement(g: dict, no: int) -> str:
    value = norm(g.get("達成度"))

    if no == 1:
        return "〇" if value in ["1", "達成"] else ""
    if no == 2:
        return "〇" if value in ["2", "一部達成"] else ""
    if no == 3:
        return "〇" if value in ["3", "未達成"] else ""

    return ""


def has_product_value(p: dict) -> bool:
    return bool(
        norm(p.get("サービス名"))
        or norm(p.get("商品名"))
        or norm(p.get("利用開始日"))
        or norm(p.get("モニタリング備考"))
    )


def fill_monitoring_sheet(wb, data: dict):
    ws = wb["レイアウト_モニタリング"]

    mapping = {
        "AC3": data.get("実施日", ""),
        "AC4": data.get("前回実施日", ""),

        "AP3": data.get("利用者名", ""),
        "AP2": data.get("利用者名カナ", ""),
        "AP4": data.get("出力氏名", ""),

        "C12": data.get("利用者名カナ", ""),
        "C13": data.get("利用者名", ""),
        "R13": data.get("性別", ""),
        "T13": data.get("生年月日", ""),
        "Y13": data.get("年齢", ""),
        "AA13": data.get("介護度", ""),
        "AE13": data.get("認定開始日", ""),
        "AK13": data.get("認定終了日", ""),

        "AC5": data.get("お話伺った人1", ""),
        "AG5": data.get("お話伺った人2", ""),
        "AJ5": data.get("お話伺った人3", ""),
        "AL5": data.get("お話伺った人その他", ""),

        "AC6": data.get("確認方法1", ""),
        "AG6": data.get("確認方法2", ""),

        "AC8": data.get("担当名", ""),

        "AP5": data.get("住所", ""),
        "AR15": data.get("住所1", ""),
        "AR16": data.get("住所2", ""),
        "AR17": data.get("住所3", ""),

        "AP6": data.get("電話番号", ""),
        "AI15": data.get("電話番号1", ""),
        "AI16": data.get("ケアマネ姓名", ""),

        "J91": data.get("専門相談員による結果", ""),
        "AA98": data.get("次回予定日", ""),

        "F85": mark_choice(data.get("身体状況の変化1", ""), ["なし", "0"]),
        "F86": mark_choice(data.get("身体状況の変化2", ""), ["あり", "1"]),
        "J85": data.get("身体状況の変化備考", ""),

        "Z85": mark_choice(data.get("ご家族状況の変化1", ""), ["なし", "0"]),
        "Z86": mark_choice(data.get("ご家族状況の変化2", ""), ["あり", "1"]),
        "AD85": data.get("ご家族状況の変化備考", ""),

        "F87": mark_choice(data.get("お気持ちの変化1", ""), ["なし", "0"]),
        "F88": mark_choice(data.get("お気持ちの変化2", ""), ["あり", "1"]),
        "J87": data.get("お気持ちの変化備考", ""),

        "Z87": mark_choice(data.get("生活状況の変化1", ""), ["なし", "0"]),
        "Z88": mark_choice(data.get("生活状況の変化2", ""), ["あり", "1"]),
        "AD87": data.get("生活状況の変化備考", ""),

        "F91": mark_choice(data.get("見直しの必要性1", ""), ["なし", "0"]),
        "F94": mark_choice(data.get("見直しの必要性2", ""), ["あり", "1"]),
    }

    for cell_ref, value in mapping.items():
        set_if_exists(ws, cell_ref, value)

    goal_rows = [20, 23, 26, 29]
    goals = data.get("福祉用具利用目標", [])

    for i, row in enumerate(goal_rows):
        g = goals[i] if i < len(goals) else DEFAULT_GOAL

        set_if_exists(ws, f"C{row}", g.get("目標", ""))
        set_if_exists(ws, f"V{row}", mark_achievement(g, 1))
        set_if_exists(ws, f"V{row + 1}", mark_achievement(g, 2))
        set_if_exists(ws, f"V{row + 2}", mark_achievement(g, 3))
        set_if_exists(ws, f"Z{row}", g.get("備考", ""))

    product_rows = [35, 41, 47, 53, 59, 65, 71, 77]
    products = data.get("商品一覧", [])

    for i, row in enumerate(product_rows):
        p = products[i] if i < len(products) else DEFAULT_PRODUCT

        if not has_product_value(p):
            continue

        item_row = row + 3

        set_if_exists(ws, f"A{row}", p.get("対応理由記号", ""))
        set_if_exists(ws, f"A{item_row}", p.get("選択制対象区分", ""))

        set_if_exists(ws, f"C{row}", p.get("サービス名", ""))
        set_if_exists(ws, f"O{row}", p.get("利用開始日", ""))
        set_if_exists(ws, f"AB{row}", p.get("モニタリング備考", ""))

        set_if_exists(
            ws,
            f"R{row}",
            mark_choice(
                p.get("使用状況の問題1", ""),
                ["なし", "0", "問題なし"]
            )
        )

        set_if_exists(
            ws,
            f"U{row}",
            mark_choice(
                p.get("点検結果1", ""),
                ["問題なし", "なし", "0"]
            )
        )

        set_if_exists(
            ws,
            f"Y{row}",
            mark_choice(
                p.get("今後の方針1", ""),
                ["継続", "1"]
            )
        )

        set_if_exists(ws, f"C{item_row}", p.get("商品名", ""))

        set_if_exists(
            ws,
            f"R{item_row}",
            mark_choice(
                p.get("使用状況の問題2", ""),
                ["あり", "1", "問題あり"]
            )
        )

        set_if_exists(
            ws,
            f"U{item_row}",
            mark_choice(
                p.get("点検結果2", ""),
                ["問題あり", "あり", "1"]
            )
        )

        set_if_exists(
            ws,
            f"Y{item_row}",
            mark_choice(
                p.get("今後の方針2", ""),
                ["再検討", "2"]
            )
        )

    return wb


@app.get("/")
def root():
    return {
        "ok": True,
        "service": "monitoring-excel-api",
    }


@app.get("/health")
def health():
    return {
        "ok": True,
        "template_exists": os.path.exists(TEMPLATE_PATH),
        "template_path": TEMPLATE_PATH,
        "version": "excel_output_choice_mark_v3",
    }


@app.post("/api/report-excel")
def report_excel(req: ExcelRequest):
    if not os.path.exists(TEMPLATE_PATH):
        raise HTTPException(
            status_code=500,
            detail="template file not found"
        )

    raw_text = (
        req.summary_json
        or req.text
        or req.transcript
        or ""
    ).strip()

    if not raw_text:
        raise HTTPException(
            status_code=400,
            detail="summary_json or text or transcript empty"
        )

    try:
        report = safe_json(raw_text)

        wb = load_workbook(TEMPLATE_PATH)
        wb = fill_monitoring_sheet(wb, report)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition":
                'attachment; filename="monitoring_report.xlsx"'
            },
        )

    except HTTPException:
        raise

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
